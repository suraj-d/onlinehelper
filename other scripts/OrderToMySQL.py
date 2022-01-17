import mysql.connector
import openpyxl
import os.path


def send_to_mysql(ws, start_row, last_row):
    mydb = mysql.connector.connect(
        host='192.168.0.2',  # '192.168.0.2'
        port='3306',  # 3306
        username='sunfashion',
        password='8632',
        database='online_database'
    )

    if not mydb:
        return {'error': "Not able to connect to database"}

    cursor = mydb.cursor()
    # file_path = r"D:\Desktop\Amazon Upload file\0. Entry in tally"
    # excel_file_name = r"OrderXML.xlsm"
    # excel_file_loc = os.path.join(file_path, excel_file_name)
    #
    # wb = openpyxl.load_workbook(excel_file_loc)
    # ws = wb['For Tally']

    # define rows and table data
    # active_cell = ws.views.sheetView[0].selection[0].activeCell
    # max_row = ws.max_row
    # max_col = ws.max_column
    # start_row = 2
    # last_row = max_row
    # entry_range = input(f'''All Data Range: {start_row}-{max_row}
    # Active Row: {active_cell}
    # Last Row: {max_row}
    # Enter Data Range: ''')

    # get data range to insert in to mysql
    try:
        # split_data = entry_range.split('-')
        # start_row = int(split_data[0])
        # last_row = int(split_data[1])
        if start_row > last_row:
            return {'error': 'Start row cannot be more than last row'}
    except ValueError:
        return {'error': 'Invalid Input'}

    def insert_order_data(tally_vch_number, date, order_id, customer_name, gst_states_id, sku_id_with_color, quantity,
                          rate, shipping,
                          cgst, sgst, igst, round_off, total, portal_name_id, warehouse_id, tally_company_id):
        args = [tally_vch_number, date, order_id, customer_name, gst_states_id, sku_id_with_color, quantity, rate,
                shipping,
                cgst, sgst, igst, round_off, total, portal_name_id, warehouse_id, tally_company_id]
        cursor.callproc('insert_order_data', args)

    def update_order_data_rate(rate, shipping, cgst, sgst, igst, round_off, total, tally_vch_number):
        args = [rate, shipping, cgst, sgst, igst,
                round_off, total, tally_vch_number]
        cursor.callproc('update_order_data_rate', args)

    new_data_count = 0
    for row in ws.iter_rows(min_row=start_row, max_row=last_row, values_only=True):
        tally_vch_number = str(row[0])
        date = row[1]
        order_id = str(row[2])
        customer_name = str(row[3])[:20]
        gst_states_id = str(row[4])
        sku_id_with_color = str(row[5])
        sku_id = str(row[6])
        quantity = row[7]
        rate = row[8]
        shipping = row[9]
        cgst = row[10]
        sgst = row[11]
        igst = row[12]
        round_off = row[13]
        total = row[14]
        portal_name_id = str(row[15])
        warehouse_id = str(row[16])
        tally_company_id = str(row[17])

        # insert query
        insert_order_data(tally_vch_number, date, order_id, customer_name, gst_states_id, sku_id_with_color, quantity,
                          rate, shipping,
                          cgst, sgst, igst, round_off, total, portal_name_id, warehouse_id, tally_company_id)

        # #update query
        # update_order_data_rate(rate, shipping, cgst, sgst,
        #                        igst, round_off, total, tally_vch_number)

        # commit
        mydb.commit()

        new_data_count += cursor.rowcount
        print(new_data_count)
        return {'number': new_data_count}


# get command
# cursor.execute(
#     f'select * from order_data')
# results = cursor.fetchall()
# for x in results:
#     print(x)

print('Done')
