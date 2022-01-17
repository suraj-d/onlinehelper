from datetime import datetime
import os.path
import openpyxl

file_path = r"D:\Desktop\Amazon Upload file\0. Entry in tally"
excel_file_name = r"ReturnXML.xlsx"
excel_file_loc = os.path.join(file_path, excel_file_name)

wb = openpyxl.load_workbook(excel_file_loc)
ws = wb['For Tally']

# rows and table data
active_cell = ws.views.sheetView[0].selection[0].activeCell
max_row = ws.max_row
max_col = ws.max_column
start_row = 2
last_row = max_row
entry_range = input(f'''All Data Range: {start_row}-{max_row}
Active Row: {active_cell}
Last Row: {max_row}
Enter Data Range: ''')

# split data range to custom start and end row
if entry_range:
    try:
        split_data = entry_range.split('-')
        start_row = int(split_data[0])
        last_row = int(split_data[1])
        if start_row > last_row:
            raise Exception
    except ValueError:
        print("Invalid Input")
        exit()
    except Exception:
        print('Start row cannot be more than last row')
        exit()


def head_xml(tally_company_id):
    return f'''<?xml version="1.0"?>
    <ENVELOPE>
    <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
    </HEADER>
    <BODY>
    <IMPORTDATA>
    <REQUESTDESC>
    <REPORTNAME>Vouchers</REPORTNAME>
    <STATICVARIABLES>
        <SVCURRENTCOMPANY>{tally_company_id}</SVCURRENTCOMPANY>
    </STATICVARIABLES>
    </REQUESTDESC>
    <REQUESTDATA>'''


def body_xml(return_date_format, tally_vch_number, return_order_id, return_type, design_name, design_number, piece, rate, shipping,
             cgst, sgst, igst, round_off, total, order_date_format, portal_name, order_gst_state, customer_name,
             current_date_format):

    return f""" <TALLYMESSAGE xmlns: UDF = "TallyUDF">
    <VOUCHER REMOTEID = ""VCHKEY = ""VCHTYPE = "Credit Note"ACTION = "Create"OBJVIEW = "Invoice Voucher View">
     <OLDAUDITENTRYIDS.LIST TYPE = "Number">
      <OLDAUDITENTRYIDS> -1 </OLDAUDITENTRYIDS>
     </OLDAUDITENTRYIDS.LIST>
     <DATE>{return_date_format}</DATE>
    <REFERENCEDATE>{return_date_format}</REFERENCEDATE>
    <VATPARTYTRANSRETURNDATE>{order_date_format}</VATPARTYTRANSRETURNDATE>
    <GUID>e5df2bc1-c84a-4a0b-99f1-f7839af473fd-00001255</GUID>
    <GSTREGISTRATIONTYPE>Consumer</GSTREGISTRATIONTYPE>
    <STATENAME>{order_gst_state}</STATENAME>
    <NARRATION>{return_type} {design_name} {order_gst_state} {return_order_id}</NARRATION>
    <COUNTRYOFRESIDENCE>India</COUNTRYOFRESIDENCE>
    <PARTYGSTIN></PARTYGSTIN>
    <PLACEOFSUPPLY>{order_gst_state}</PLACEOFSUPPLY>
    <CLASSNAME>OnlineReturn</CLASSNAME>
     <PARTYNAME>{portal_name}</PARTYNAME>
     <PARTYLEDGERNAME>{portal_name}</PARTYLEDGERNAME>
     <VATPARTYTRANSRETURNNUMBER>{return_order_id}</VATPARTYTRANSRETURNNUMBER>
     <URDORIGINALSALEVALUE> Lesser than or equal to 2.5 lakhs </URDORIGINALSALEVALUE>
     <GSTNATUREOFRETURN> 01-Sales Return </GSTNATUREOFRETURN>
     <VOUCHERTYPENAME> Credit Note </VOUCHERTYPENAME>
     <REFERENCE>{return_order_id}</REFERENCE>
     <VOUCHERNUMBER>{tally_vch_number}</VOUCHERNUMBER>
     <BASICBASEPARTYNAME>{portal_name}</BASICBASEPARTYNAME>
     <PERSISTEDVIEW> Invoice Voucher View </PERSISTEDVIEW>
     <CONSIGNEEGSTIN> </CONSIGNEEGSTIN>
     <BASICBUYERNAME>{customer_name}</BASICBUYERNAME>
     <BASICDATETIMEOFREMOVAL>{current_date_format}</BASICDATETIMEOFREMOVAL>
     <VCHGSTCLASS/>
     <CONSIGNEESTATENAME>{order_gst_state}</CONSIGNEESTATENAME>
     <ENTEREDBY> admin </ENTEREDBY>
     <DIFFACTUALQTY> No </DIFFACTUALQTY>
     <ISMSTFROMSYNC> No </ISMSTFROMSYNC>
     <ASORIGINAL> No </ASORIGINAL>
     <AUDITED> No </AUDITED>
     <FORJOBCOSTING> No </FORJOBCOSTING>
     <ISOPTIONAL> No </ISOPTIONAL>
     <EFFECTIVEDATE>{return_date_format}</EFFECTIVEDATE>
     <USEFOREXCISE> No </USEFOREXCISE>
     <ISFORJOBWORKIN> No </ISFORJOBWORKIN>
     <ALLOWCONSUMPTION> No </ALLOWCONSUMPTION>
     <USEFORINTEREST> No </USEFORINTEREST>
     <USEFORGAINLOSS> No </USEFORGAINLOSS>
     <USEFORGODOWNTRANSFER> No </USEFORGODOWNTRANSFER>
     <USEFORCOMPOUND> No </USEFORCOMPOUND>
     <USEFORSERVICETAX> No </USEFORSERVICETAX>
     <ISDELETED> No </ISDELETED>
     <ISONHOLD> No </ISONHOLD>
     <ISBOENOTAPPLICABLE> No </ISBOENOTAPPLICABLE>
     <ISEXCISEVOUCHER> No </ISEXCISEVOUCHER>
     <EXCISETAXOVERRIDE> No </EXCISETAXOVERRIDE>
     <USEFORTAXUNITTRANSFER> No </USEFORTAXUNITTRANSFER>
     <IGNOREPOSVALIDATION> No </IGNOREPOSVALIDATION>
     <EXCISEOPENING> No </EXCISEOPENING>
     <USEFORFINALPRODUCTION> No </USEFORFINALPRODUCTION>
     <ISTDSOVERRIDDEN> No </ISTDSOVERRIDDEN>
     <ISTCSOVERRIDDEN> No </ISTCSOVERRIDDEN>
     <ISTDSTCSCASHVCH> No </ISTDSTCSCASHVCH>
     <INCLUDEADVPYMTVCH> No </INCLUDEADVPYMTVCH>
     <ISSUBWORKSCONTRACT> No </ISSUBWORKSCONTRACT>
     <ISVATOVERRIDDEN> No </ISVATOVERRIDDEN>
     <IGNOREORIGVCHDATE> No </IGNOREORIGVCHDATE>
     <ISVATPAIDATCUSTOMS> No </ISVATPAIDATCUSTOMS>
     <ISDECLAREDTOCUSTOMS> No </ISDECLAREDTOCUSTOMS>
     <ISSERVICETAXOVERRIDDEN> No </ISSERVICETAXOVERRIDDEN>
     <ISISDVOUCHER> No </ISISDVOUCHER>
     <ISEXCISEOVERRIDDEN> No </ISEXCISEOVERRIDDEN>
     <ISEXCISESUPPLYVCH> No </ISEXCISESUPPLYVCH>
     <ISGSTOVERRIDDEN> No </ISGSTOVERRIDDEN>
     <GSTNOTEXPORTED> No </GSTNOTEXPORTED>
     <IGNOREGSTINVALIDATION> No </IGNOREGSTINVALIDATION>
     <ISGSTREFUND> No </ISGSTREFUND>
     <ISGSTSECSEVENAPPLICABLE> No </ISGSTSECSEVENAPPLICABLE>
     <ISVATPRINCIPALACCOUNT> No </ISVATPRINCIPALACCOUNT>
     <ISSHIPPINGWITHINSTATE> No </ISSHIPPINGWITHINSTATE>
     <ISOVERSEASTOURISTTRANS> No </ISOVERSEASTOURISTTRANS>
     <ISDESIGNATEDZONEPARTY> No </ISDESIGNATEDZONEPARTY>
     <ISCANCELLED> No </ISCANCELLED>
     <HASCASHFLOW> No </HASCASHFLOW>
     <ISPOSTDATED> No </ISPOSTDATED>
     <USETRACKINGNUMBER> No </USETRACKINGNUMBER>
     <ISINVOICE> Yes </ISINVOICE>
     <MFGJOURNAL> No </MFGJOURNAL>
     <HASDISCOUNTS> No </HASDISCOUNTS>
     <ASPAYSLIP> No </ASPAYSLIP>
     <ISCOSTCENTRE> No </ISCOSTCENTRE>
     <ISSTXNONREALIZEDVCH> No </ISSTXNONREALIZEDVCH>
     <ISEXCISEMANUFACTURERON> No </ISEXCISEMANUFACTURERON>
     <ISBLANKCHEQUE> No </ISBLANKCHEQUE>
     <ISVOID> No </ISVOID>
     <ORDERLINESTATUS> No </ORDERLINESTATUS>
     <VATISAGNSTCANCSALES> No </VATISAGNSTCANCSALES>
     <VATISPURCEXEMPTED> No </VATISPURCEXEMPTED>
     <ISVATRESTAXINVOICE> No </ISVATRESTAXINVOICE>
     <VATISASSESABLECALCVCH> No </VATISASSESABLECALCVCH>
     <ISVATDUTYPAID> Yes </ISVATDUTYPAID>
     <ISDELIVERYSAMEASCONSIGNEE> No </ISDELIVERYSAMEASCONSIGNEE>
     <ISDISPATCHSAMEASCONSIGNOR> No </ISDISPATCHSAMEASCONSIGNOR>
     <CHANGEVCHMODE> No </CHANGEVCHMODE>
     <ALTERID> 8791 </ALTERID>
     <MASTERID> 4693 </MASTERID>
     <VOUCHERKEY> 189884799123592 </VOUCHERKEY>
     <INVENTORYENTRIES.LIST>
      <STOCKITEMNAME>{design_number}</STOCKITEMNAME>
      <ISDEEMEDPOSITIVE> Yes </ISDEEMEDPOSITIVE>
      <ISLASTDEEMEDPOSITIVE> Yes </ISLASTDEEMEDPOSITIVE>
      <ISAUTONEGATE> No </ISAUTONEGATE>
      <ISCUSTOMSCLEARANCE> No </ISCUSTOMSCLEARANCE>
      <ISTRACKCOMPONENT> No </ISTRACKCOMPONENT>
      <ISTRACKPRODUCTION> No </ISTRACKPRODUCTION>
      <ISPRIMARYITEM> No </ISPRIMARYITEM>
      <ISSCRAP> No </ISSCRAP>
      <RATE>{rate} Pcs </RATE>
      <AMOUNT>-{rate}</AMOUNT>
      <ACTUALQTY>{piece} Pcs </ACTUALQTY>
      <BILLEDQTY>{piece} Pcs </BILLEDQTY>
         <BATCHALLOCATIONS.LIST>
              <GODOWNNAME> Ready Stock Godown </GODOWNNAME>
              <BATCHNAME> Primary Batch </BATCHNAME>
              <DESTINATIONGODOWNNAME> Ready Stock Godown </DESTINATIONGODOWNNAME>
              <DYNAMICCSTISCLEARED> No </DYNAMICCSTISCLEARED>
              <AMOUNT>-{rate}</AMOUNT>
              <ACTUALQTY>{piece}Pcs</ACTUALQTY>
              <BILLEDQTY>{piece}Pcs</BILLEDQTY>
      </BATCHALLOCATIONS.LIST>
      <ACCOUNTINGALLOCATIONS.LIST>
              <OLDAUDITENTRYIDS.LIST TYPE = "Number">
               <OLDAUDITENTRYIDS> -1 </OLDAUDITENTRYIDS>
              </OLDAUDITENTRYIDS.LIST>
          <LEDGERNAME>Online Return</LEDGERNAME>
          <CLASSRATE>100.00000</CLASSRATE>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <LEDGERFROMITEM>No</LEDGERFROMITEM>
          <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
          <ISPARTYLEDGER>No</ISPARTYLEDGER>
          <ISLASTDEEMEDPOSITIVE>Yes</ISLASTDEEMEDPOSITIVE>
          <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
          <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
          <AMOUNT>-{rate}</AMOUNT>
      </ACCOUNTINGALLOCATIONS.LIST>
     </INVENTORYENTRIES.LIST>
    <LEDGERENTRIES.LIST>
     <OLDAUDITENTRYIDS.LIST TYPE="Number">
       <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
      </OLDAUDITENTRYIDS.LIST>
      <LEDGERNAME>{portal_name}</LEDGERNAME>
      <GSTCLASS/>
      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>Yes</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <AMOUNT>{total}</AMOUNT>
      <BILLALLOCATIONS.LIST>
       <NAME>{return_order_id}</NAME>
       <BILLTYPE>Agst Ref</BILLTYPE>
       <TDSDEDUCTEEISSPECIALRATE>No</TDSDEDUCTEEISSPECIALRATE>
       <AMOUNT>{total}</AMOUNT>
      </BILLALLOCATIONS.LIST>
     </LEDGERENTRIES.LIST>
    
    <LEDGERENTRIES.LIST>
      <OLDAUDITENTRYIDS.LIST TYPE="Number">
          <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
      </OLDAUDITENTRYIDS.LIST>
      <LEDGERNAME>Shipping Charge Collected</LEDGERNAME>
      <METHODTYPE>As User Defined Value</METHODTYPE>
      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>Yes</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>No</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <AMOUNT>-{shipping}</AMOUNT>
      <VATEXPAMOUNT>-{shipping}</VATEXPAMOUNT>
  </LEDGERENTRIES.LIST>

    <LEDGERENTRIES.LIST>
         <OLDAUDITENTRYIDS.LIST TYPE="Number">
              <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
          </OLDAUDITENTRYIDS.LIST>
      <ROUNDTYPE/>
      <LEDGERNAME>CGST</LEDGERNAME>
      <METHODTYPE>GST</METHODTYPE>
      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>Yes</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>No</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>Yes</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <AMOUNT>-{cgst}</AMOUNT>
      <VATEXPAMOUNT>-{cgst}</VATEXPAMOUNT>
     </LEDGERENTRIES.LIST>

     <LEDGERENTRIES.LIST>
         <OLDAUDITENTRYIDS.LIST TYPE="Number">
              <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
          </OLDAUDITENTRYIDS.LIST>
      <ROUNDTYPE/>
      <LEDGERNAME>SGST</LEDGERNAME>
      <METHODTYPE>GST</METHODTYPE>
      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>Yes</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>No</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>Yes</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <AMOUNT>-{sgst}</AMOUNT>
      <VATEXPAMOUNT>-{sgst}</VATEXPAMOUNT>
     </LEDGERENTRIES.LIST>

      <LEDGERENTRIES.LIST>
         <OLDAUDITENTRYIDS.LIST TYPE="Number">
              <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
          </OLDAUDITENTRYIDS.LIST>
      <ROUNDTYPE/>
      <LEDGERNAME>IGST</LEDGERNAME>
      <METHODTYPE>GST</METHODTYPE>
      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>Yes</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>No</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>Yes</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <AMOUNT>-{igst}</AMOUNT>
      <VATEXPAMOUNT>-{igst}</VATEXPAMOUNT>
     </LEDGERENTRIES.LIST>

     
    <LEDGERENTRIES.LIST>
      <OLDAUDITENTRYIDS.LIST TYPE="Number">
          <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
      </OLDAUDITENTRYIDS.LIST>
      <ROUNDTYPE>Normal Rounding</ROUNDTYPE>
      <LEDGERNAME>Rounded Off</LEDGERNAME>
      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
      <LEDGERFROMITEM>No</LEDGERFROMITEM>
      <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
      <ISPARTYLEDGER>No</ISPARTYLEDGER>
      <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
      <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
      <ISCAPVATNOTCLAIMED>No</ISCAPVATNOTCLAIMED>
      <ROUNDLIMIT> 1</ROUNDLIMIT>
      <AMOUNT>{round_off}</AMOUNT>
      <VATEXPAMOUNT>{round_off}</VATEXPAMOUNT>
  </LEDGERENTRIES.LIST>
    </VOUCHER>
   </TALLYMESSAGE>
"""


def tail_xml(tally_company_id):
    return f"""<TALLYMESSAGE xmlns:UDF="TallyUDF">
        <COMPANY>
            <REMOTECMPINFO.LIST MERGE="Yes">
               <NAME>e5df2bc1-c84a-4a0b-99f1-f7839af473fd</NAME>
               <REMOTECMPNAME>{tally_company_id}</REMOTECMPNAME>
               <REMOTECMPSTATE>Gujarat</REMOTECMPSTATE>
           </REMOTECMPINFO.LIST>
       </COMPANY>
    </TALLYMESSAGE>
    
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
       <COMPANY>
           <REMOTECMPINFO.LIST MERGE="Yes">
               <NAME>e5df2bc1-c84a-4a0b-99f1-f7839af473fd</NAME>
               <REMOTECMPNAME>{tally_company_id}</REMOTECMPNAME>
               <REMOTECMPSTATE>Gujarat</REMOTECMPSTATE>
           </REMOTECMPINFO.LIST>
       </COMPANY>
    </TALLYMESSAGE>    
    </REQUESTDATA>
    </IMPORTDATA>
    </BODY>
    </ENVELOPE>
    """


data_number = 0
tally_company_id = ws.cell(row=start_row, column=19).value
xml_string = head_xml(tally_company_id)
for row in ws.iter_rows(min_row=start_row, max_row=last_row, max_col=max_col, values_only=True):
    return_date = row[0]
    tally_vch_number = row[1]
    return_order_id = row[2]
    return_type = row[3]
    design_name = row[4]
    design_number = row[5]
    piece = row[6]
    rate = row[7]
    shipping = row[8]
    cgst = row[9]
    sgst = row[10]
    igst = row[11]
    round_off = row[12]*-1
    total = row[13]
    order_date = row[14]
    portal_name = row[15]
    order_gst_state = row[16]
    customer_name = row[17]
    tally_company = row[18]
    order_data_id = row[19]
    return_initiated_id = row[20]
    penalty = row[21]

    return_date_format = datetime.strftime(return_date, '%Y%m%d')
    order_date_format = datetime.strftime(order_date, '%Y%m%d')
    current_date_format = datetime.strftime(
        datetime.now(), '%d-%b-%Y at %H:%M')

    xml_string += body_xml(return_date_format, tally_vch_number, return_order_id, return_type, design_name, design_number, piece, rate, shipping,
                           cgst, sgst, igst, round_off, total, order_date_format, portal_name, order_gst_state, customer_name,
                           current_date_format)

    data_number += 1
    print(f'{data_number} done')


xml_string += tail_xml(tally_company_id)

# create text file
txt_file_name = f'onlineXML.txt'
txt_file_loc = os.path.join(file_path, txt_file_name)
f = open(txt_file_loc, 'w+')
f.write(xml_string)
f.close()
print(f'{data_number} File Generated')
