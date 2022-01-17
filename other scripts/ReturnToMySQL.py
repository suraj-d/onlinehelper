import mysql.connector
import openpyxl
import os.path

mydb = mysql.connector.connect(
    host='192.168.0.2',  # '192.168.0.2'
    port='3306',  # 3306
    username='sunfashion',
    password='8632',
    database='online_database'
)

if not mydb:
    print("Not able to connect to database")
    exit()

cursor = mydb.cursor()
file_path = r"D:\Desktop\Amazon Upload file\0. Entry in tally"
excel_file_name = r"ReturnXML.xlsx"
excel_file_loc = os.path.join(file_path, excel_file_name)

wb = openpyxl.load_workbook(excel_file_loc)
ws = wb['For Tally']

# define rows and table data
active_cell = ws.views.sheetView[0].selection[0].activeCell
max_row = ws.max_row
max_col = ws.max_column
start_row = 2
last_row = max_row
entry_range = input(f'''All Data Range: {start_row}-{max_row}
Active Row: {active_cell}
Last Row: {max_row}
Enter Data Range: ''')

# get data range to inster in to mysql
if entry_range:
    try:
        split_data = entry_range.split('-')
        start_row = int(split_data[0])
        last_row = int(split_data[1])
        if start_row > last_row:
            raise Exception
    except ValueError:
        print("Invalid Input")
        exit()
    except IndexError:
        start_row = 2
        print(f"Input should be {start_row}-{max_row}")
        exit()
    except Exception:
        print('Start row cannot be more than last row')
        exit()




# get command
# cursor.execute(
#     f'select * from order_data')
# results = cursor.fetchall()
# for x in results:
#     print(x)

print('Done')
