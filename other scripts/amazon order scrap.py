from openpyxl import load_workbook
import json

# update file location and sheet name
wb = load_workbook(
    r'S:\4. Accounting\0. Entry in tally\Order entry check.xlsx')
ws = wb['Sheet1']  # case sentitive

# amazon scrap
head = r'''{"_id":"amazon_price","startUrl":'''
# url = r'https://sellercentral.amazon.in/orders-v3/order/'
# tail = r""","selectors":[{"delay":0,"id":"order_id","multiple":false,"parentSelectors":["_root"],"regex":"","selector":".a-span12 span.a-text-bold","type":"SelectorText"},{"delay":0,"id":"sku","multiple":false,"parentSelectors":["orderDetaiBox"],"regex":"","selector":"div.product-name-column-word-wrap-break-all:nth-of-type(3) div","type":"SelectorText"},{"delay":0,"id":"rate","multiple":false,"parentSelectors":["orderDetaiBox"],"regex":"","selector":"td.a-text-right span","type":"SelectorText"},{"delay":0,"id":"orderDetaiBox","multiple":true,"parentSelectors":["_root"],"selector":".a-spacing-large tbody tr","type":"SelectorElement"},{"delay":0,"id":"buyer_name","multiple":false,"parentSelectors":["_root"],"regex":"","selector":"span div > span:nth-of-type(1)","type":"SelectorText"},{"delay":0,"id":"gst_state","multiple":false,"parentSelectors":["_root"],"regex":"","selector":"span span:nth-of-type(5)","type":"SelectorText"}]}"""
mid = []

# flipkart scrap
url = u"""https://seller.flipkart.com/index.html#dashboard/my-orders?serviceProfile=seller-fulfilled&shipmentType=easy-ship&orderState=shipments_delivered&orderItemId="""
tail = r""","selectors":[{"id":"click","parentSelectors":["_root"],"type":"SelectorElementClick","clickElementSelector":"td.clickable","clickElementUniquenessType":"uniqueText","clickType":"clickOnce","delay":2000,"discardInitialElements":"do-not-discard","multiple":false,"selector":"td.clickable"},{"id":"name","parentSelectors":["_root"],"type":"SelectorText","selector":"div.styles__ItemWrapperVertical-sc-9qxfse-1:nth-of-type(2) div:nth-of-type(2)","multiple":false,"delay":0,"regex":""},{"id":"state","parentSelectors":["_root"],"type":"SelectorText","selector":".styles__ItemWrapperVertical-sc-9qxfse-1 div:nth-of-type(6)","multiple":false,"delay":0,"regex":""},{"id":"box","parentSelectors":["_root"],"type":"SelectorElement","selector":"div.styles__OrderItemContainer-sc-9qxfse-6","multiple":true,"delay":0},{"id":"orderid","parentSelectors":["box"],"type":"SelectorText","selector":"div.styles__ItemDetails-sc-1bxatwx-6:nth-of-type(2) div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""},{"id":"sku","parentSelectors":["box"],"type":"SelectorText","selector":"div.styles__ItemDetails-sc-1bxatwx-6:nth-of-type(1) div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""},{"id":"rate","parentSelectors":["box"],"type":"SelectorText","selector":"div:nth-of-type(1) div:nth-of-type(6) div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""}]}"""

start_row = 1
last_row = 38

for row in ws.iter_rows(values_only=True, min_row=start_row):
    # print(f"https://sellercentral.amazon.in/orders-v3/order/{row[0]}")
    mid.append(f'{url}{row[0]}')

full_json = f'{head}{json.dumps(mid)}{tail}'
f = open('amazonOrder.txt', 'w+')
f.write(full_json)
f.close()
print("Done")


def get_json_string(portal_name):
    head = None
    mid = []
    url = None
    tail = None
    if portal_name.lower() == 'amazon':
        head = r'{"_id":"amazon_price","startUrl":'
        url = r'https://sellercentral.amazon.in/orders-v3/order/'
        tail = r""","selectors":[{"delay":0,"id":"order_id","multiple":false,"parentSelectors":["_root"],"regex":"",
        "selector":".a-span12 span.a-text-bold","type":"SelectorText"},{"delay":0,"id":"sku","multiple":false,
        "parentSelectors":["orderDetaiBox"],"regex":"",
        "selector":"div.product-name-column-word-wrap-break-all:nth-of-type(3) div","type":"SelectorText"},
        {"delay":0,"id":"rate","multiple":false,"parentSelectors":["orderDetaiBox"],"regex":"",
        "selector":"td.a-text-right span","type":"SelectorText"},{"delay":0,"id":"orderDetaiBox","multiple":true,
        "parentSelectors":["_root"],"selector":".a-spacing-large tbody tr","type":"SelectorElement"},{"delay":0,
        "id":"buyer_name","multiple":false,"parentSelectors":["_root"],"regex":"","selector":"span div > 
        span:nth-of-type(1)","type":"SelectorText"},{"delay":0,"id":"gst_state","multiple":false,"parentSelectors":[
        "_root"],"regex":"","selector":"span span:nth-of-type(5)","type":"SelectorText"}]} """
    elif portal_name.lower() == 'flipkart':
        head = r'{"_id":"fk_order_with_click_delete","startUrl":'
        url = u"""https://seller.flipkart.com/index.html#dashboard/my-orders?serviceProfile=seller-fulfilled
        &shipmentType=easy-ship&orderState=shipments_delivered&orderItemId= """
        tail = r""","selectors":[{"id":"click","parentSelectors":["_root"],"type":"SelectorElementClick",
        "clickElementSelector":"td.clickable","clickElementUniquenessType":"uniqueText","clickType":"clickOnce",
        "delay":2000,"discardInitialElements":"do-not-discard","multiple":false,"selector":"td.clickable"},
        {"id":"name","parentSelectors":["_root"],"type":"SelectorText",
        "selector":"div.styles__ItemWrapperVertical-sc-9qxfse-1:nth-of-type(2) div:nth-of-type(2)","multiple":false,
        "delay":0,"regex":""},{"id":"state","parentSelectors":["_root"],"type":"SelectorText",
        "selector":".styles__ItemWrapperVertical-sc-9qxfse-1 div:nth-of-type(6)","multiple":false,"delay":0,
        "regex":""},{"id":"box","parentSelectors":["_root"],"type":"SelectorElement",
        "selector":"div.styles__OrderItemContainer-sc-9qxfse-6","multiple":true,"delay":0},{"id":"orderid",
        "parentSelectors":["box"],"type":"SelectorText","selector":"div.styles__ItemDetails-sc-1bxatwx-6:nth-of-type(
        2) div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""},{"id":"sku","parentSelectors":[
        "box"],"type":"SelectorText","selector":"div.styles__ItemDetails-sc-1bxatwx-6:nth-of-type(1) 
        div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""},{"id":"rate","parentSelectors":[
        "box"],"type":"SelectorText","selector":"div:nth-of-type(1) div:nth-of-type(6) 
        div.styles__Value-sc-1bxatwx-14","multiple":false,"delay":0,"regex":""}]} """

