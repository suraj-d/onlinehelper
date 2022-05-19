from datetime import datetime

from openpyxl import load_workbook

from src.xmlFormats import head_xml, payment_body_1_xml, payment_bank_xml, payment_advance_fee_xml, \
    payment_general_exp_xml, payment_unavl_xml, payment_order_data_1_xml, payment_order_data_2_xml, \
    payment_order_data_3_xml, tail_xml

print('start')


def get_payment_xml_file(excel_file_path, sheet_name):
    """
    :param excel_file_path:
    :param sheet_name:
    :return dict: xml_string or error in case of error
    """
    xml_string = ""
    wb = load_workbook(excel_file_path)
    ws = wb[sheet_name]
    print('excel sheet loaded')
    # Particular data
    ws_particular_table = ws.tables['particularTable']
    particular_data_range = ws[ws_particular_table.ref]
    particular_data = []
    try:
        for cell in particular_data_range:
            ledger = cell[0].value
            debit = cell[1].value
            credit = cell[2].value

            # print(f'{ledger}, {debit}, {credit}')
            if ledger not in ("Particular", 'Total'):
                if debit not in (None, 0, ""):
                    amount = round(abs(debit), 2) * -1
                    deemed_positive = "Yes"
                elif credit not in (None, 0, ""):
                    amount = round(abs(credit), 2)
                    deemed_positive = "No"
                else:
                    amount = 0
                    ledger = ""
                    deemed_positive = ""

                particular_data.append([ledger, amount, deemed_positive])

    except Exception as e:
        return {'error': e}
    # print(particular_date)

    # assign variables
    tally_company = "Sun Fashion And Lifestyle"
    vch_no = ws['b3'].value
    ref_no = ws['d20'].value
    entry_date = datetime.strftime(ws['f3'].value, '%Y%m%d')
    advance_fee_ledger = particular_data[0][0]
    advance_fee_amnt = particular_data[0][1]
    advance_fee_deemed_positive = particular_data[0][2]
    tcs_igst_ledger = particular_data[1][0]
    tcs_igst_amt = particular_data[1][1]
    tcs_igst_deemed_positive = particular_data[1][2]
    tcs_sgst_ledger = particular_data[2][0]
    tcs_sgst_amt = particular_data[2][1]
    tcs_sgst_deemed_positive = particular_data[2][2]
    tcs_cgst_ledger = particular_data[3][0]
    tcs_cgst_amt = particular_data[3][1]
    tcs_cgst_deemed_positive = particular_data[3][2]
    tds_ledger = particular_data[4][0]
    tds_amt = particular_data[4][1]
    tds_deemed_positive = particular_data[4][2]
    unavbl_ledger = particular_data[5][0]
    unavbl_amt_dr = particular_data[5][1]
    unavbl_deemed_positive_dr = particular_data[5][2]
    unavbl_amt_cr = particular_data[6][1]
    unavbl_deemed_positive_cr = particular_data[6][2]
    unavbl_agstName = ws['f11'].value
    reimb_ledger = particular_data[7][0]
    reimb_amt = particular_data[7][1]
    reimb_deemed_positive = particular_data[7][2]
    misc_adj_ledger = particular_data[8][0]
    misc_adj_amt = particular_data[8][1]
    misc_adj_deemed_positive = particular_data[8][2]
    bank_ledger = particular_data[9][0]
    bank_amt = particular_data[9][1]
    bank_deemed_positive = particular_data[9][2]
    portal_name = particular_data[10][0]
    portal_amt = particular_data[10][1]
    portal_deemed_positive = particular_data[10][2]
    narration = f'{portal_name} settlement id: {ref_no}'

    xml_string += head_xml(tally_company_id=tally_company)

    xml_string += payment_body_1_xml(entry_date=entry_date, narration=narration, portal_name=portal_name, ref_no=ref_no,
                                     vch_no=vch_no)
    if bank_ledger != "":
        xml_string += payment_bank_xml(ledger=bank_ledger, amount=bank_amt,
                                       deemed_positive=bank_deemed_positive, entry_date=entry_date,
                                       portal_name=portal_name)
    if advance_fee_ledger != "":
        xml_string += payment_advance_fee_xml(ledger=advance_fee_ledger, amount=advance_fee_amnt,
                                              deemed_positive=advance_fee_deemed_positive, ref_no=ref_no)
    if tcs_igst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_igst_ledger, amount=tcs_igst_amt,
                                              deemed_positive=tcs_igst_deemed_positive)
    if tcs_cgst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_cgst_ledger, amount=tcs_cgst_amt,
                                              deemed_positive=tcs_cgst_deemed_positive)
    if tcs_sgst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_sgst_ledger, amount=tcs_sgst_amt,
                                              deemed_positive=tcs_sgst_deemed_positive)
    if tds_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tds_ledger, amount=tds_amt,
                                              deemed_positive=tds_deemed_positive)
    if unavbl_ledger != "":
        xml_string += payment_unavl_xml(ledger=unavbl_ledger, amount=unavbl_amt_dr,
                                        deemed_positive=unavbl_deemed_positive_dr, ref_no=unavbl_agstName)

        xml_string += payment_unavl_xml(ledger=unavbl_ledger, amount=unavbl_amt_cr,
                                        deemed_positive=unavbl_deemed_positive_cr, ref_no=ref_no)
    if reimb_ledger != "":
        xml_string += payment_general_exp_xml(ledger=reimb_ledger, amount=reimb_amt,
                                              deemed_positive=reimb_deemed_positive)
    if misc_adj_ledger != "":
        xml_string += payment_general_exp_xml(ledger=misc_adj_ledger, amount=misc_adj_amt,
                                              deemed_positive=misc_adj_deemed_positive)
    if portal_name != "":
        xml_string += payment_order_data_1_xml(ledger=portal_name, amount=portal_amt,
                                               deemed_positive=portal_deemed_positive)

    # Order ids and amount details
    ws_order_data_table = ws.tables['orderData']
    order_data_range = ws[ws_order_data_table.ref]

    try:
        for cell in order_data_range:
            order_id = cell[0].value
            amount = cell[1].value
            bill_type = cell[2].value
            if order_id is None:
                break
            if order_id != "orderID":
                xml_string += payment_order_data_2_xml(order_id=order_id, amount=amount, bill_type=bill_type)

    except Exception as e:
        return {'error': e}

    xml_string += payment_order_data_3_xml()
    xml_string += tail_xml(tally_company_id=tally_company)

    return {"xml_string": xml_string}


excel = r"S:\4. Accounting\0. Entry in tally\PaymentXML.xlsx"
excel2 = r'C:\Users\sunfashionLap\Desktop\Book1.xlsx'
sheet = "PaymentXML"
print('loading wb')
try:
    wb = load_workbook(excel2)
    print(wb)
except Exception as e:
    print(e)
# get_payment_xml_file(excel, sheet)
print('done')
