from datetime import datetime

import win32com.client as win32
import sys
from os import path, startfile
from csv import reader
from subprocess import Popen
from openpyxl import load_workbook, Workbook
from PyQt5.QtWidgets import QFileDialog, QApplication
from PyQt5.QtCore import QSettings
from PyQt5 import uic

from src.CommanFunction import create_xlsx_file

if __name__ == "__main__":
    gui_file = "../gui/createOrderSheet.ui"
else:
    gui_file = "gui/createOrderSheet.ui"
Ui_order_sheet, baseclass = uic.loadUiType(gui_file)


class OrderSheetWindow(baseclass):

    def __init__(self):
        super(OrderSheetWindow, self).__init__()
        uic.loadUi(gui_file, self)

        # function value
        self.file_path = None
        self.portal_selected = None
        self.save_file = None

        # Disable button
        self.browse_button.setDisabled(True)
        self.generate_sheet_button.setDisabled(True)
        self.open_folder_button.setDisabled(True)

        try:
            # button connection
            self.portal_combo_box.addItems(["Select Company", "Amazon", "Flipkart", "Meesho", "Shopee"])
            self.portal_combo_box.currentIndexChanged.connect(self.selectionchange)  # get value on selection change
            self.file_path_input.textChanged.connect(self.validate_gen_name)
            self.gen_name_input.textChanged.connect(self.validate_gen_name)

            self.browse_button.clicked.connect(self.browse_files)
            self.generate_sheet_button.clicked.connect(self.create_order_sheet)
            self.open_folder_button.clicked.connect(self.open_folder)
        except Exception as e:
            self.result_input.setPlainText(e)

        self.show()

    def selectionchange(self, i):
        """
        get combo box value on select
        """
        try:
            self.portal_selected = self.portal_combo_box.currentText()
            self.browse_button.setDisabled(False)
        except Exception as e:
            print(e)

    def browse_files(self):
        try:
            setting_last_file_path = QSettings("Order Helper", "LastSetting")  # path in regedit
            last_open_file_path = setting_last_file_path.value("order_sheet_path")  # get value

            if self.portal_selected.lower() == "amazon":
                allowed_file_format = "Text File (*.txt);;All Files (*.*)"
            else:
                allowed_file_format = "Excel File (*.xls *.xlsx *.csv);;All Files (*.*)"

            filepath = QFileDialog.getOpenFileName(self, 'Open Order Sheet File', last_open_file_path,
                                                   allowed_file_format)

            if filepath[0] != "":
                self.portal_combo_box.setDisabled(True)
                self.file_path = path.normpath(filepath[0])  # backslash in folder path text
                self.file_path_input.setText(self.file_path)  # set text in line edit
                folder_path = path.split(self.file_path_input.text())[0]  # get folder path
                setting_last_file_path.setValue("order_sheet_path", folder_path)
        except Exception as e:
            print(e)

    def validate_gen_name(self):
        self.open_folder_button.setDisabled(True)
        if self.gen_name_input.text() and self.file_path_input.text():
            self.portal_combo_box.setDisabled(True)
            self.generate_sheet_button.setDisabled(False)
        elif not self.gen_name_input.text() and not self.file_path_input.text():
            self.portal_combo_box.setDisabled(False)
        else:
            self.generate_sheet_button.setDisabled(True)

    def create_order_sheet(self):
        try:
            folder_path = path.split(self.file_path_input.text())[0]
            self.save_file = path.join(folder_path, f"{self.gen_name_input.text()}.xlsx")
            base_sheet_name = "Original Data"

            # portal wise data selection
            order_id_row = None
            buyer_name_row = None
            state_row = None
            amount_row = None
            sku_row = None
            qty_row = None
            start_row = None
            converted_file_path = None

            # zero start excel row
            if self.portal_selected.lower() == 'amazon':
                order_id_row = 0
                buyer_name_row = 8
                state_row = 22
                amount_row = 26
                sku_row = 10
                qty_row = 15
                start_row = 2

                # get original data base excel file path
                converted_file_path = csv_to_xlsx(self.file_path, self.save_file, base_sheet_name)

            elif self.portal_selected.lower() == 'flipkart':
                # reader_list = reader(f)
                order_id_row = 2
                buyer_name_row = 20
                state_row = 25
                amount_row = 19
                sku_row = 8
                qty_row = 18
                start_row = 2

                # get original data base excel file path
                converted_file_path = csv_to_xlsx(self.file_path, self.save_file, base_sheet_name)

            elif self.portal_selected.lower() == 'meesho':
                order_id_row = 2
                buyer_name_row = 5
                state_row = 7
                amount_row = 15
                sku_row = 9
                qty_row = 11
                start_row = 4

                # # get original data base excel file path
                converted_file_path = xls_to_xlsx(self.file_path, self.save_file, base_sheet_name)
            elif self.portal_selected.lower() == 'shopee':
                order_id_row = 1
                buyer_name_row = 8
                state_row = 9
                amount_row = 4
                sku_row = 6
                qty_row = 5
                start_row = 2

                original_sheet_name = "Original Order Data"
                copy_sheet = xls_to_xlsx(self.file_path,self.save_file, original_sheet_name)
                converted_file_path = shopee_original_data(copy_sheet.get("save_path"), original_sheet_name)




            # load saved original data base file
            save_loc = converted_file_path.get('save_path')
            wb = load_workbook(save_loc)
            sku_tally_sheet(wb, start_row, sku_row, order_id_row, buyer_name_row, qty_row, state_row, amount_row,
                            save_loc)

            self.open_folder_button.setDisabled(False)
            self.result_input.setPlainText(f"File Generated on {datetime.now().strftime('%Y-%m-%d')} "
                                           f"at {datetime.now().strftime('%H:%M:%S')}")
            if self.open_check_box.isChecked():
                self.auto_open_file()
        except Exception as e:
            print(f"create order sheet: {e}")
            self.result_input.setPlainText(e)

    def open_folder(self):
        try:
            Popen(fr'explorer /select,{self.file_path}')
        except Exception as e:
            print(e)

    def auto_open_file(self):
        startfile(self.save_file)

    def closeEvent(self, event):
        try:
            pass
        except Exception as e:
            print(e)


# FUNCTIONS
def xls_to_xlsx(file_path, save_path, sheet_name):
    # convert xls to xlsx for meesho
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file_path)
    wb.SaveAs(save_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()

    # set wb and sheet name
    wb = load_workbook(save_path)
    ws1 = wb.active
    ws1.title = sheet_name  # change sheet name

    wb.save(save_path)

    return {'save_path': save_path}


def csv_to_xlsx(file_path, save_path, sheet_name):
    f = open(file_path, 'r', encoding="utf8")

    if 'txt' in file_path:
        reader_list = reader(f, delimiter="\t") # amazon format
    else:
        reader_list = reader(f) # flipkart format

    # # set wb and sheet name
    # wb = Workbook()  # load_workbook(file_path)
    # ws1 = wb.active
    # ws1.title = sheet_name  # change sheet name
    #
    # # add data to excel
    # for row in reader_list:
    #     ws1.append(row)
    #
    # wb.save(save_path)
    create_xlsx_file(reader_list, save_path, sheet_name)

    return {'save_path': save_path}


def sku_tally_sheet(wb, start_row, sku_row, order_id_row, buyer_name_row, qty_row, state_row, amount_row, save_path):
    wb.create_sheet(index=1, title="SKU List")  # new sheet created
    ws_data = wb["Original Data"]  # get data form original data sheet
    ws2 = wb["SKU List"]  # set new sheet

    sku_list = [["Sku", "Qty"]]  # sku and qty list with header
    for_tally_list = [["Order id", "Buyer Name", "State", "Sku", "Amount", "Qty"]]
    for row in ws_data.iter_rows(min_row=start_row, max_row=ws_data.max_row, values_only=True):
        """
        loop through original data to get sku and qty to ship 
        row start at '0'
        """
        sku = row[sku_row]
        if row[qty_row] is not None:
            qty = int(row[qty_row])
        else:
            qty = row[qty_row]

        order_id = row[order_id_row]
        buyer_name = row[buyer_name_row]
        state = row[state_row]
        amount = row[amount_row]

        '''check if sku exist, if yes add qty else append to list'''
        for sku_data in sku_list:
            if sku_data[0] == sku:
                sku_data[1] += qty
                break
        else:
            sku_list.append([sku, qty])

        for_tally_list.append([order_id, buyer_name, state, sku, amount, qty])

    # add data to sku sheet
    for row in sku_list:
        ws2.append(row)

    # for tally sheet
    wb.create_sheet(index=2, title="For Tally")
    ws3 = wb['For Tally']

    # add data to tally sheet
    for row in for_tally_list:
        ws3.append(row)

    # save all
    wb.save(save_path)

    return {'save_path': save_path}


def shopee_original_data(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws_data = wb[sheet_name]  # get data form original data sheet

    wb.create_sheet(index=1, title='Original Data')
    ws1 = wb['Original Data']
    new_original_data = [['Tracking ID', 'Order ID', 'Product Name', 'Variation Name',
                          'Price', 'Quantity', 'Sku', 'Parent sku', 'Buyer Name', 'State']]
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
        tracking_id = row[0]
        order_id = row[1]
        product_detail = row[2].split(";")
        product_name = product_detail[0].split(':')[1].strip()
        variation_name = product_detail[1].split(':')[1].strip()
        price = product_detail[2].split(' ')[3]
        qty = product_detail[3].split(':')[1].strip()
        sku = product_detail[4].split(':')[1].strip()
        parent_sku = product_detail[5].split(':')[1].strip()
        buyer_name = "Shopee cust"
        state = "Check order pdf"

        new_original_data.append([tracking_id, order_id, product_name, variation_name, price, qty, sku, parent_sku, buyer_name, state])

    for row in new_original_data:
        ws1.append(row)

    # save all
    wb.save(file_path)

    return {'save_path': file_path}


if __name__ == '__main__':
    app = QApplication(sys.argv)
    windows = OrderSheetWindow()
    sys.exit(app.exec_())
