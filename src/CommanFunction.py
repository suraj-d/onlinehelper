from openpyxl import load_workbook, Workbook
from os import path
from mysql.connector import connect


# read excel file and return row data
def read_excel_sheet(excel_file_path, sheet_name):
    """
    :param excel_file_path: excel sheet full path
    :param sheet_name: sheet name to read
    :return dict: active_cell, max_row, max_col, start_row, last_row, worksheet detail
    """
    try:
        print("Reading Excel Sheet...")
        # excel_file_loc = path.join(folder_path, file_name)
        wb = load_workbook(excel_file_path)
        ws = wb[sheet_name]

        # rows and table data
        return {
            'active_cell': ws.views.sheetView[0].selection[0].activeCell,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'start_row': 2,
            'last_row': ws.max_row,
            'ws': ws
            # 'last_row': max_row,
        }
    except Exception as e:
        return {'error': str(e)}


def get_excel_sheet_name(excel_file_path):
    """
    :param excel_file_path:
    :return: sheet_name list
    """
    try:
        wb = load_workbook(excel_file_path)
        return {'sheet_name': wb.sheetnames}
    except Exception as e:
        return {'error': str(e)}


def create_xlsx_file(content, save_path, sheet_name="Sheet 1"):
    """
    :param content: content to write in excel sheet
    :param save_path: save location
    :param sheet_name: saved sheet name
    :return save file location: new saved file location
    """
    wb = Workbook()  # load_workbook(file_path)
    ws = wb.active
    ws.title = sheet_name  # change sheet name

    # add data to excel
    for row in content:
        ws.append(row)

    wb.save(save_path)

    return {'save_path': save_path}


# create text file and return text file location
def create_text_file(content: str, folder_path, txt_file_name):
    """
    :param content: content text string
    :param folder_path: folder full path
    :param txt_file_name: file name
    :return: string of save file path
    """
    txt_file_loc = path.join(folder_path, f'{txt_file_name}.txt')
    f = open(txt_file_loc, 'w+')
    f.write(content)
    f.close()
    return path.normpath(txt_file_loc)


# get sql connection and return mysqldb or error
def get_sql_connection():
    """

    :return: mysqldb (database) or error
    """
    try:
        mysqldb = connect(
            host='sunserver',  # '192.168.0.2'
            port='3306',  # 3306
            username='sunfashion',
            password='8632',
            database='online_database'
        )
        # print(mysqldb)
    except Exception as e:
        return {'error': e}
    else:
        return {"mysqldb": mysqldb}

# get data from table
def get_table_data(query, args=None):
    """

    :param query: select query
    :param args: any arguments to pass in select query
    :return: "data", multi tuple
    """
    sql_validate = get_sql_connection()
    if 'error' in sql_validate:
        error = sql_validate.get("error")
        return {'error': error}

    mysqldb = sql_validate.get('mysqldb')
    cursor = mysqldb.cursor()

    cursor.execute(query, args)
    data = cursor.fetchall()
    # print(data)

    return {'data': data}


