import sys
from datetime import datetime
from os import path

from PyQt5 import uic
from PyQt5.QtCore import QSettings
from PyQt5.QtWidgets import QApplication, QFileDialog
from openpyxl import load_workbook
from pyperclip import copy

from src.CommanFunction import create_text_file, read_excel_sheet, validate_user_data
from src.xmlFormats import head_xml, tail_xml, stock_item_name_start, stock_item_name_end, \
    stock_item_component_list_start, stock_item_component_item, stock_item_component_list_end

if __name__ == "__main__":
    gui_file = "../gui/createStockItemXML.ui"
else:
    gui_file = "gui/createStockItemXML.ui"

Ui_order_form, baseClass = uic.loadUiType(gui_file)


class StockItemWindow(baseClass):

    def __init__(self):
        super(StockItemWindow, self).__init__()
        self.ws = None
        uic.loadUi(gui_file, self)

        self.browse_button.clicked.connect(self.browse_files)
        self.read_excel_button.clicked.connect(self.read_excel)
        self.generate_xml_button.clicked.connect(self.generate_xml_file)
        self.copy_button.clicked.connect(self.copy_path)

        self.sheet_name_input.setText("Stock Item")
        self.show()

    def browse_files(self):
        setting_last_file_path = QSettings("Order Helper", "LastSetting")  # path in regedit
        reg_edit_name = "payment_xml_sheet_path"
        last_open_file_path = setting_last_file_path.value(reg_edit_name)  # get value

        # open dialog box
        filepath = QFileDialog.getOpenFileName(self, 'Open File', last_open_file_path, "Excel File (*.xls *.xlsx)")

        if filepath[0] != "":
            file_path = path.normpath(filepath[0])

            self.excel_file_path_input.setText(file_path)  # set value in line edit

            # split full path in folder name and file name from line edit
            folder_path, file_name = path.split(self.excel_file_path_input.text())

            self.file_name_output.setText(str(file_name))

            setting_last_file_path.setValue("order_xml_sheet_path", folder_path)  # save path to regedit

            self.read_excel_button.setEnabled(True)

    def read_excel(self):
        self.excel_file_path_input.setDisabled(True)
        self.sheet_name_input.setDisabled(True)
        self.file_name_output.setDisabled(True)

        excel_file_path = self.excel_file_path_input.text()
        sheet_name = self.sheet_name_input.text()

        # read excel sheet
        excel_data = read_excel_sheet(excel_file_path, sheet_name)

        if "error" not in excel_data:
            excel_data_string: str = f'WorkSheet: {excel_data.get("ws")} loaded'

            self.ws = excel_data.get('ws')

            # activate button
            self.generate_xml_button.setDisabled(False)
        else:
            excel_data_string: str = f"Error: {excel_data.get('error')}"

        self.excel_data_output.setText(str(excel_data_string))

    def generate_xml_file(self):
        try:

                # create xml string
                xml_data = get_stock_item_xml_file(self.ws)

                # write text file
                if 'error' in xml_data:
                    print(xml_data.get("error"))
                    xml_msg: str = f'Error: {xml_data.get("error")}'
                else:
                    xml_string = xml_data.get('xml_string')
                    output_file_name = xml_data.get("file_name")
                    folder_path = path.split(self.excel_file_path_input.text())[0]
                    text_location = create_text_file(xml_string, folder_path, output_file_name)
                    self.save_path_input.setText(str(text_location))
                    self.copy_button.setDisabled(False)
                    xml_msg: str = f'{xml_data.get("number")} stock item xml generated\n' \
                                   f'File Location: {text_location}'
                self.xml_data_output.setText(str(xml_msg))
        except Exception as e:
            xml_msg = f'generate_xml_file error: {e}'
            self.xml_data_output.setText(str(xml_msg))

    def copy_path(self):
        copy(self.save_path_input.text())
        self.copy_button.setText("Copied")


#####################
### MAIN FUNCTION ###
#####################


def get_stock_item_xml_file(ws):
    stock_item_start = False
    component_item_start = False
    check_stock_item_name = ""
    data_number = 0
    xml_string = head_xml()

    stock_item_table = ws.tables["stock_item_table"]
    stock_item_table_range = ws[stock_item_table.ref]

    component_item_table = ws.tables["component_item_table"]
    component_item_table_range = ws[component_item_table.ref]

    #
    # for row in ws.iter_rows(min_row=start_row, max_row=last_row, max_col=ws.max_column, values_only=True):
    #     stock_item_type = str(row[0])
    #     stock_item_name = str(row[1])
    #     parent = str(row[2])
    #     category = str(row[3])
    #     base_unit = str(row[4])
    #     alternate_unit = str(row[5])
    #     base_qty = int(row[6])
    #     alternate_qty = int(row[7])
    #     component_list_name = str(row[8])
    #     component_base_qty = str(row[9])
    #     component_stock_item_name = str(row[10])
    #     component_godown_name = str(row[11])
    #     component_stock_item_qty = str(row[12])
    #
    #     if stock_item_type == "Stock Item":
    #         stock_item_start = True
    #         check_stock_item_name = stock_item_name
    #         data_number += 1
    #         xml_string += stock_item_name_start(stock_item_name, parent, category, base_unit, alternate_unit, base_qty,
    #                                             alternate_qty)
    #     elif stock_item_type == "Component List":
    #         component_item_start = True
    #         xml_string += stock_item_component_list_start(component_list_name, component_base_qty)
    #     elif stock_item_type == "Component Item":
    #         xml_string += stock_item_component_item(component_stock_item_name, component_godown_name,
    #                                                 component_stock_item_qty)
    #
    #     if component_item_start:
    #         xml_string += stock_item_component_list_end()
    #
    #     if stock_item_start:
    #         xml_string += stock_item_name_end()
    #
    #     print(f'{data_number} done')
    #     print(row)

    for stock_item_cell in stock_item_table_range:
        stock_item_name = stock_item_cell[0].value
        parent = stock_item_cell[1].value
        category = stock_item_cell[2].value
        base_unit = stock_item_cell[3].value
        alternate_unit = stock_item_cell[4].value
        base_qty = stock_item_cell[5].value
        alternate_qty = stock_item_cell[6].value
        component_list_name = stock_item_cell[7].value
        component_base_qty = stock_item_cell[8].value
        component_base_unit = stock_item_cell[9].value

        # add stock item name start xml
        if stock_item_name == "Stock Item Name":
            continue

        if stock_item_name != "":
            data_number += 1
            xml_string += stock_item_name_start(stock_item_name, parent, category, base_unit, alternate_unit, base_qty,
                                                alternate_qty)
        # add component list start xml
        if component_list_name != "":
            xml_string += stock_item_component_list_start(component_list_name, component_base_qty, component_base_unit)

        print(stock_item_name)
        # check component table for data with stock_item_name
        for component_item_cell in component_item_table_range:
            component_stock_item_ref = component_item_cell[0].value
            component_stock_item_name = component_item_cell[1].value
            component_godown_name = component_item_cell[2].value
            component_stock_item_qty = component_item_cell[3].value
            component_stock_item_unit = component_item_cell[4].value

            if component_stock_item_ref == "Stock Item Name":
                continue
            print(component_stock_item_name)
            if component_stock_item_ref == stock_item_name:
                xml_string += stock_item_component_item(component_stock_item_name, component_godown_name,
                                                        component_stock_item_qty, component_stock_item_unit)
            else:
                continue
        # add component list end xml
        if component_list_name != "":
            xml_string += stock_item_component_list_end()

        # add stock item name end xml
        if stock_item_name != "":
            xml_string += stock_item_name_end()

    xml_string += tail_xml()

    return {'xml_string': xml_string,
            'number': data_number,
            'file_name': "stock item XML"}


def get_payment_xml_file(excel_file_path, sheet_name):
    """
    :param excel_file_path:
    :param sheet_name:
    :return dict: xml_string or error in case of error
    """
    xml_string = ""
    wb = load_workbook(excel_file_path)
    ws = wb[sheet_name]
    print('excel sheet loaded')
    # Particular data
    ws_particular_table = ws.tables['particularTable']
    particular_data_range = ws[ws_particular_table.ref]
    particular_data = []
    try:
        for cell in particular_data_range:
            ledger = cell[0].value
            debit = cell[1].value
            credit = cell[2].value

            # print(f'{ledger}, {debit}, {credit}')
            if ledger not in ("Particular", 'Total'):
                if debit not in (None, 0, ""):
                    amount = round(abs(debit), 2) * -1
                    deemed_positive = "Yes"
                elif credit not in (None, 0, ""):
                    amount = round(abs(credit), 2)
                    deemed_positive = "No"
                else:
                    amount = 0
                    ledger = ""
                    deemed_positive = ""

                particular_data.append([ledger, amount, deemed_positive])

    except Exception as e:
        return {'error': e}
    # print(particular_date)

    # assign variables
    tally_company = "Sun Fashion And Lifestyle"
    vch_no = ws['b3'].value
    ref_no = ws['d20'].value
    entry_date = datetime.strftime(ws['f3'].value, '%Y%m%d')
    advance_fee_ledger = particular_data[0][0]
    advance_fee_amnt = particular_data[0][1]
    advance_fee_deemed_positive = particular_data[0][2]
    tcs_igst_ledger = particular_data[1][0]
    tcs_igst_amt = particular_data[1][1]
    tcs_igst_deemed_positive = particular_data[1][2]
    tcs_sgst_ledger = particular_data[2][0]
    tcs_sgst_amt = particular_data[2][1]
    tcs_sgst_deemed_positive = particular_data[2][2]
    tcs_cgst_ledger = particular_data[3][0]
    tcs_cgst_amt = particular_data[3][1]
    tcs_cgst_deemed_positive = particular_data[3][2]
    tds_ledger = particular_data[4][0]
    tds_amt = particular_data[4][1]
    tds_deemed_positive = particular_data[4][2]
    unavbl_ledger = particular_data[5][0]
    unavbl_amt_dr = particular_data[5][1]
    unavbl_deemed_positive_dr = particular_data[5][2]
    unavbl_amt_cr = particular_data[6][1]
    unavbl_deemed_positive_cr = particular_data[6][2]
    unavbl_agstName = ws['f11'].value
    reimb_ledger = particular_data[7][0]
    reimb_amt = particular_data[7][1]
    reimb_deemed_positive = particular_data[7][2]
    misc_adj_ledger = particular_data[8][0]
    misc_adj_amt = particular_data[8][1]
    misc_adj_deemed_positive = particular_data[8][2]
    bank_ledger = particular_data[9][0]
    bank_amt = particular_data[9][1]
    bank_deemed_positive = particular_data[9][2]
    portal_name = particular_data[10][0]
    portal_amt = particular_data[10][1]
    portal_deemed_positive = particular_data[10][2]
    narration = f'{portal_name} settlement id: {ref_no}'

    xml_string += head_xml(tally_company_id=tally_company)

    xml_string += payment_body_1_xml(entry_date=entry_date, narration=narration, portal_name=portal_name, ref_no=ref_no,
                                     vch_no=vch_no)
    if bank_ledger != "":
        xml_string += payment_bank_xml(ledger=bank_ledger, amount=bank_amt,
                                       deemed_positive=bank_deemed_positive, entry_date=entry_date,
                                       portal_name=portal_name)
    if advance_fee_ledger != "":
        xml_string += payment_advance_fee_xml(ledger=advance_fee_ledger, amount=advance_fee_amnt,
                                              deemed_positive=advance_fee_deemed_positive, ref_no=ref_no)
    if tcs_igst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_igst_ledger, amount=tcs_igst_amt,
                                              deemed_positive=tcs_igst_deemed_positive)
    if tcs_cgst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_cgst_ledger, amount=tcs_cgst_amt,
                                              deemed_positive=tcs_cgst_deemed_positive)
    if tcs_sgst_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tcs_sgst_ledger, amount=tcs_sgst_amt,
                                              deemed_positive=tcs_sgst_deemed_positive)
    if tds_ledger != "":
        xml_string += payment_general_exp_xml(ledger=tds_ledger, amount=tds_amt,
                                              deemed_positive=tds_deemed_positive)
    if unavbl_ledger != "":
        xml_string += payment_unavl_xml(ledger=unavbl_ledger, amount=unavbl_amt_dr,
                                        deemed_positive=unavbl_deemed_positive_dr, ref_no=unavbl_agstName)

        xml_string += payment_unavl_xml(ledger=unavbl_ledger, amount=unavbl_amt_cr,
                                        deemed_positive=unavbl_deemed_positive_cr, ref_no=ref_no)
    if reimb_ledger != "":
        xml_string += payment_general_exp_xml(ledger=reimb_ledger, amount=reimb_amt,
                                              deemed_positive=reimb_deemed_positive)
    if misc_adj_ledger != "":
        xml_string += payment_general_exp_xml(ledger=misc_adj_ledger, amount=misc_adj_amt,
                                              deemed_positive=misc_adj_deemed_positive)
    if portal_name != "":
        xml_string += payment_order_data_1_xml(ledger=portal_name, amount=portal_amt,
                                               deemed_positive=portal_deemed_positive)

    # Order ids and amount details
    ws_order_data_table = ws.tables['orderData']
    order_data_range = ws[ws_order_data_table.ref]

    try:
        for cell in order_data_range:
            order_id = cell[0].value
            amount = cell[1].value
            bill_type = cell[2].value
            if order_id is None:
                break
            if order_id != "orderID":
                xml_string += payment_order_data_2_xml(order_id=order_id, amount=amount, bill_type=bill_type)

    except Exception as e:
        return {'error': e}

    xml_string += payment_order_data_3_xml()
    xml_string += tail_xml(tally_company_id=tally_company)

    return {"xml_string": xml_string}


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = StockItemWindow()

    sys.exit(app.exec_())
