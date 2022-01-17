from openpyxl import load_workbook

wb = load_workbook(r'S:\4. Accounting\0. Entry in tally\PaymentXML.xlsx', data_only=True)
ws = wb['PaymentXML']
ws_table = ws.tables['orderData']

try:

    print('name:'+ ws_table.name)
    print(ws_table.ref)
    a = ws_table.ref
    print(type(a))
    for row in ws.iter_rows(range_string = a).value:
        print(row)
except Exception as e:
    print(e)

# start from here
for i in xrange(0, 100):
    ...
    for j in xrange(0, 100):
        ...
    ws.cell(row=i, column=j)
